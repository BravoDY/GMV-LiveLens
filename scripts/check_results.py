import openpyxl

wb = openpyxl.load_workbook(r"D:\MyDownloads\阿发_已匹配.xlsx")
ws = wb["收货省市"]

print("=== 抽样查看 N列结果 ===")
print(f"{'Row':<6} {'M(数仓-市)':<40} {'N(城市编码)':<60}")
for r in [2, 3, 4, 5, 10, 11, 14, 15, 16, 17, 18, 19, 20, 21, 22, 28, 30, 32, 34, 36, 39, 40, 41, 42, 44, 45, 50, 60, 70, 80, 100, 200, 300, 500, 800, 1000, 1200, 1268]:
    if r > ws.max_row:
        break
    m = ws.cell(row=r, column=13).value
    n = ws.cell(row=r, column=14).value
    print(f"{r:<6} {str(m)[:38]:<40} {str(n)[:58]:<60}")

print()
print("=== 未找到的M列值分布 (前50个) ===")
from collections import Counter
not_found = []
for r in range(2, ws.max_row + 1):
    n = ws.cell(row=r, column=14).value
    m = ws.cell(row=r, column=13).value
    if n and isinstance(n, str) and "未找到" in n:
        not_found.append(str(m).strip() if m else "空")

counter = Counter(not_found)
for city, cnt in counter.most_common(50):
    print(f"  [{city}] x{cnt}")
print(f"  ... 共 {len(counter)} 种不同值")
