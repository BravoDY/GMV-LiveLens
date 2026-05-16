import openpyxl

wb = openpyxl.load_workbook(r"D:\MyDownloads\阿发.xlsx")
ws = wb["收货省市"]

c_cities = set()
d_map = {}
for r in range(2, ws.max_row + 1):
    c = ws.cell(row=r, column=3).value
    d = ws.cell(row=r, column=4).value
    if c and d:
        c_cities.add(str(c).strip())
        d_map[str(c).strip()] = str(d).strip()

m_cities = set()
for r in range(2, ws.max_row + 1):
    m = ws.cell(row=r, column=13).value
    if m:
        m_cities.add(str(m).strip())

exact_match = c_cities & m_cities
only_c = c_cities - m_cities
only_m = m_cities - c_cities

print(f"Total C cities: {len(c_cities)}")
print(f"Total M cities: {len(m_cities)}")
print(f"Exact matches: {len(exact_match)}")
print(f"Only in C: {len(only_c)}")
print(f"Only in M: {len(only_m)}")
print()
print("=== Only in M (need fuzzy match) ===")
for city in sorted(only_m)[:80]:
    print(f"  [{city}]")
if len(only_m) > 80:
    print(f"  ... and {len(only_m) - 80} more")
print()
print("=== Only in C ===")
for city in sorted(only_c)[:80]:
    print(f"  [{city}] -> code: {d_map.get(city, '?')}")
if len(only_c) > 80:
    print(f"  ... and {len(only_c) - 80} more")
