import openpyxl
import re
from copy import copy

SRC = r"D:\MyDownloads\阿发.xlsx"
DST = r"D:\MyDownloads\阿发.xlsx"

wb = openpyxl.load_workbook(SRC)
ws = wb["收货省市"]

print(f"总行数: {ws.max_row}")

# ============================================================
# Step 1: 从 C→D 构建城市→编码映射表
# ============================================================
city_to_code = {}
for r in range(2, ws.max_row + 1):
    c = ws.cell(row=r, column=3).value
    d = ws.cell(row=r, column=4).value
    if c is not None and d is not None:
        c = str(c).strip()
        d = str(d).strip()
        if c and d:
            city_to_code[c] = d

print(f"C→D 映射条目: {len(city_to_code)}")


# ============================================================
# Step 2: 构建规范化名称辅助函数
# ============================================================
def normalize(name):
    """去掉常见行政区划后缀，返回规范化名称"""
    suffixes = [
        "傣族景颇族自治州", "蒙古族藏族自治州", "布依族苗族自治州",
        "土家族苗族自治州", "哈尼族彝族自治州", "壮族苗族自治州",
        "藏族羌族自治州", "苗族侗族自治州", "傈僳族自治州",
        "朝鲜族自治州", "黎族苗族自治县", "黎族自治县",
        "藏族自治州", "苗族自治州", "回族自治州", "彝族自治州",
        "蒙古自治州",
        "自治州", "地区", "市", "县", "区",
    ]
    result = name
    for sfx in suffixes:
        if result.endswith(sfx) and result != sfx:
            result = result[:-len(sfx)]
            break
    return result


def extract_core(name):
    """提取城市核心名，去除所有后缀"""
    result = name
    # 反复去除后缀直到不变
    changed = True
    while changed:
        changed = False
        suffixes = [
            "傣族景颇族自治州", "蒙古族藏族自治州", "布依族苗族自治州",
            "土家族苗族自治州", "哈尼族彝族自治州", "壮族苗族自治州",
            "藏族羌族自治州", "苗族侗族自治州", "傈僳族自治州",
            "朝鲜族自治州", "黎族苗族自治县", "黎族自治县",
            "藏族自治州", "苗族自治州", "回族自治州", "彝族自治州",
            "蒙古自治州", "自治州", "地区", "市", "县", "区", "州",
        ]
        for sfx in suffixes:
            if result.endswith(sfx) and result != sfx:
                result = result[:-len