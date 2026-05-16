import openpyxl

wb = openpyxl.load_workbook(r"D:\MyDownloads\阿发.xlsx")
ws = wb["收货省市"]

d_map = {}
for r in range(2, ws.max_row + 1):
    c = ws.cell(row=r, column=3).value
    d = ws.cell(row=r, column=4).value
    if c and d:
        d_map[str(c).strip()] = str(d).strip()

print("=== ALL C column cities with codes ===")
for name, code in sorted(d_map.items()):
    print(f"  [{name}] -> {code}")

print()
m_cities = set()
for r in range(2, ws.max_row + 1):
    m = ws.cell(row=r, column=13).value
    if m:
        m_cities.add(str(m).strip())

only_m = m_cities - set(d_map.keys())
print(f"\n=== ALL Only in M ({len(only_m)}) ===")
for city in sorted(only_m):
    print(f"  [{city}]")
