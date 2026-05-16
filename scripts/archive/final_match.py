"""
最终匹配脚本: 使用完整 GB/T 2260 编码库,覆盖所有未匹配城市
"""
import openpyxl
from gbt2260_data import GB_T2260_CITY, ALIAS_MAP, SPECIAL_REASONS

SRC = r"D:\MyDownloads\阿发.xlsx"
OUT = r"D:\MyDownloads\阿发_已匹配_v2.xlsx"

wb = openpyxl.load_workbook(SRC)
ws = wb["收货省市"]

def match_city_code(m_value):
    """返回 (code|None, reason|None)"""
    if m_value is None:
        return None, "M列为空"
    m = str(m_value).strip()
    if not m:
        return None, "M列为空"

    # 1. 特殊原因
    if m in SPECIAL_REASONS:
        return None, SPECIAL_REASONS[m]

    # 2. 精确匹配 GB/T 2260 标准名称
    if m in GB_T2260_CITY:
        return GB_T2260_CITY[m], None

    # 3. 别名映射
    if m in ALIAS_MAP:
        std_name = ALIAS_MAP[m]
        if std_name in GB_T2260_CITY:
            return GB_T2260_CITY[std_name], None

    # 4. 标准化后缀尝试
    suffixes_full = [
        "彝族自治州", "藏族自治州", "苗族自治州", "壮族自治州",
        "回族自治州", "蒙古自治州", "傣族自治州", "傈僳族自治州",
        "朝鲜族自治州", "土家族苗族自治州", "布依族苗族自治州",
        "哈尼族彝族自治州", "白族自治州", "自治州",
        "地区", "市", "州", "盟", "县", "区", "林区", "新区",
    ]

    # 尝试: M是简写(如XX州) → 找标准全称(如XX自治州)
    for full_suffix in suffixes_full:
        if m.endswith(full_suffix):
            break
    else:
        for full_suffix in ["自治州", "地区", "市", "州", "盟", "县"]:
            test_name = m + full_suffix
            if test_name in GB_T2260_CITY and not m.endswith(full_suffix):
                return GB_T2260_CITY[test_name], None

    # 5. 区域别名(海南县级市等,已是正确名称直接找)
    # 上述步骤已覆盖大部分

    # 6. 大陆城市但C列有编码的就用C列编码
    # (由外层逻辑处理,此处只做GB/T 2260匹配)

    return None, f"在GB/T2260标准库中未找到: {m}"


# 获取原始C→D映射作为兜底
c_d_map = {}
for r in range(2, ws.max_row + 1):
    c = ws.cell(row=r, column=3).value
    d = ws.cell(row=r, column=4).value
    if c and d:
        c_d_map[str(c).strip()] = str(d).strip()

stats = {"exact_c": 0, "gbt_exact": 0, "gbt_alias": 0, "gbt_suffix": 0,
         "special": 0, "not_found": 0, "empty": 0}

for r in range(2, ws.max_row + 1):
    m_val = ws.cell(row=r, column=13).value
    m_str = str(m_val).strip() if m_val else ""

    # 先尝试C列精确匹配
    if m_str in c_d_map:
        ws.cell(row=r, column=14).value = c_d_map[m_str]
        stats["exact_c"] += 1
        continue

    # GB/T 2260 匹配
    code, reason = match_city_code(m_val)

    if code is not None:
        ws.cell(row=r, column=14).value = code
        if m_str in ALIAS_MAP:
            stats["gbt_alias"] += 1
        elif m_str in GB_T2260_CITY:
            stats["gbt_exact"] += 1
        else:
            stats["gbt_suffix"] += 1
    elif reason:
        ws.cell(row=r, column=14).value = reason
        if "M列为空" in reason or "M列为null" in reason:
            stats["empty"] += 1
        elif reason in SPECIAL_REASONS.values():
            stats["special"] += 1
        else:
            stats["not_found"] += 1
    else:
        ws.cell(row=r, column=14).value = "未知原因未匹配"

wb.save(OUT)
print(f"保存成功: {OUT}")
print(f"\n=== 匹配统计 ===")
print(f"C列精确匹配:    {stats['exact_c']}")
print(f"GB/T 2260 精确: {stats['gbt_exact']}")
print(f"GB/T 2260 别名: {stats['gbt_alias']}")
print(f"GB/T 2260 后缀: {stats['gbt_suffix']}")
print(f"特殊原因:       {stats['special']}")
print(f"未找到:         {stats['not_found']}")
print(f"空值:           {stats['empty']}")
total = sum(stats.values())
print(f"\n总计: {total} / {ws.max_row - 1}")
