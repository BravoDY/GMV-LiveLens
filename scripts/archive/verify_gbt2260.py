"""
第一步: 验证已匹配编码是否是 GB/T 2260 身份证行政区划代码
"""
import openpyxl

wb = openpyxl.load_workbook(r"D:\MyDownloads\阿发_已匹配.xlsx")
ws = wb["收货省市"]

# GB/T 2260 省份代码前缀范围
PROVINCE_PREFIXES = {
    "11": "北京市", "12": "天津市", "13": "河北省", "14": "山西省",
    "15": "内蒙古", "21": "辽宁省", "22": "吉林省", "23": "黑龙江省",
    "31": "上海市", "32": "江苏省", "33": "浙江省", "34": "安徽省",
    "35": "福建省", "36": "江西省", "37": "山东省",
    "41": "河南省", "42": "湖北省", "43": "湖南省", "44": "广东省",
    "45": "广西", "46": "海南省",
    "50": "重庆市", "51": "四川省", "52": "贵州省", "53": "云南省",
    "54": "西藏",
    "61": "陕西省", "62": "甘肃省", "63": "青海省", "64": "宁夏",
    "65": "新疆",
    "71": "台湾", "81": "香港", "82": "澳门",
}

def is_valid_gbt2260(code):
    """检查编码是否符合 GB/T 2260 格式"""
    if not code or not isinstance(code, str):
        return False
    if not code.isdigit():
        return False
    if len(code) != 6:
        return False
    prefix = code[:2]
    if prefix not in PROVINCE_PREFIXES:
        return False
    return True

# 收集所有已填入的编码
codes_found = {}
for r in range(2, ws.max_row + 1):
    c_name = ws.cell(row=r, column=3).value
    c_code = ws.cell(row=r, column=4).value
    n_val = ws.cell(row=r, column=14).value
    if n_val and isinstance(n_val, str) and n_val.isdigit():
        m_name = ws.cell(row=r, column=13).value
        codes_found[str(m_name).strip() if m_name else ""] = n_val

print("=== 已匹配编码验证 ===")
all_valid = True
for m_name, code in list(codes_found.items())[:30]:
    valid = is_valid_gbt2260(code)
    province = PROVINCE_PREFIXES.get(code[:2], "?") if valid else "?"
    flag = "✓" if valid else "✗"
    print(f"  {flag} [{m_name}] → {code} ({province})")
    if not valid:
        all_valid = False

print(f"\n总计已匹配: {len(codes_found)} 条")
print(f"格式验证: {'全部通过' if all_valid else '存在异常编码'}")

# 检查是否有非6位数字异常
anomalies = []
for m_name, code in codes_found.items():
    if not is_valid_gbt2260(code):
        anomalies.append((m_name, code))
if anomalies:
    print(f"\n异常编码 ({len(anomalies)}):")
    for name, code in anomalies:
        print(f"  [{name}] → {code}")
