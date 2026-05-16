import openpyxl
import re
from copy import copy

SRC = r"D:\MyDownloads\阿发.xlsx"
OUT = r"D:\MyDownloads\阿发_已匹配.xlsx"

wb = openpyxl.load_workbook(SRC)
ws = wb["收货省市"]

d_map = {}
for r in range(2, ws.max_row + 1):
    c = ws.cell(row=r, column=3).value
    d = ws.cell(row=r, column=4).value
    if c and d:
        d_map[str(c).strip()] = str(d).strip()


def strip_suffix(name):
    """去掉 市/地区/州/盟/县/区/林区/新区 等后缀, 返回 (core, suffix)"""
    suffixes = [
        "彝族自治州", "藏族自治州", "苗族自治州", "壮族自治州",
        "回族自治州", "蒙古自治州", "傣族自治州", "傈僳族自治州",
        "朝鲜族自治州", "土家族苗族自治州", "布依族苗族自治州",
        "哈尼族彝族自治州", "白族自治州", "自治州",
        "地区", "市", "州", "盟", "县", "区", "林区", "新区",
    ]
    for sfx in suffixes:
        if name.endswith(sfx):
            return name[:-len(sfx)], sfx
    return name, ""


def smart_match(m_value, d_map):
    """返回 (code|None, reason|None)"""
    if m_value is None:
        return None, "M列为空"
    m = str(m_value).strip()
    if not m:
        return None, "M列为空"
    if m == "null":
        return None, "M列为null"

    # 1. 精确匹配
    if m in d_map:
        return d_map[m], None

    # 2. 特殊映射表 (C列中没有完整名称,但有对应的)
    special_map = {
        # 中国香港 / 中国澳门 → 多个子区域, 有歧义
        "中国香港": ("歧义", "C列中香港分:香港岛(810100)/九龙(810200)/新界(810300), 无法确定"),
        "中国澳门": ("歧义", "C列中澳门分:澳门半岛(820100)/离岛(820200), 无法确定"),
        # 台湾
        "台湾": (None, "C列中无台湾对应编码"),
        "台北市": (None, "C列中无台湾对应编码"),
        # 国外
        "加拿大": (None, "国外地址,不在中国城市编码中"),
        "日本": (None, "国外地址,不在中国城市编码中"),
        "美国": (None, "国外地址,不在中国城市编码中"),
        "英国": (None, "国外地址,不在中国城市编码中"),
        "新加坡": (None, "国外地址,不在中国城市编码中"),
        "泰国": (None, "国外地址,不在中国城市编码中"),
        "越南": (None, "国外地址,不在中国城市编码中"),
        "马来西亚": (None, "国外地址,不在中国城市编码中"),
        # 雄安新区
        "雄安新区": (None, "C列中无雄安新区对应编码"),
    }
    if m in special_map:
        return special_map[m]

    # 3. 去掉后缀后匹配 (处理 "凉山州"→"凉山彝族自治州", "伊犁州"→"伊犁哈萨克自治州")
    core, sfx = strip_suffix(m)
    if sfx:
        # 在C列中找同名核心但后缀不同的
        candidates = []
        for c_name in d_map:
            c_core, c_sfx = strip_suffix(c_name)
            if c_core == core and c_name != m:
                candidates.append((c_name, d_map[c_name]))
        if len(candidates) == 1:
            return candidates[0][1], None
        if len(candidates) > 1:
            names = ", ".join(f"{n}({c})" for n, c in candidates)
            return "歧义", f"多个候选: {names}"

    # 4. 模糊匹配: 检查M是否是C列中某个城市的简称 (去掉 "市/地区/州" 后前缀包含)
    candidates = []
    for c_name, code in d_map.items():
        c_core, _ = strip_suffix(c_name)
        if c_core == m:
            candidates.append((c_name, code))
    if len(candidates) == 1:
        return candidates[0][1], None
    if len(candidates) > 1:
        names = ", ".join(f"{n}({c})" for n, c in candidates)
        return "歧义", f"多个候选: {names}"

    # 5. M列有些是"XX地区"但C列只有"XX市" (如"吐鲁番地区"→"吐鲁番市")
    #    但如果C列根本没有"吐鲁番"相关, 则匹配不上
    #    再试: 如果M是"XX地区", 但C中有"XX市", 则匹配
    if m.endswith("地区"):
        base = m[:-2]
        # 找 C 中 base + 市
        for c_name, code in d_map.items():
            if c_name == base + "市":
                return code, None
        # 找 C 中 base + 地区 (C列本身就有地区)
        for c_name, code in d_map.items():
            if c_name == m:
                return code, None

    # 6. 无法匹配
    return None, f"在C列{len(d_map)}个城市中未找到匹配项"


stats = {"exact": 0, "fuzzy": 0, "ambiguous": 0, "not_found": 0, "empty": 0}

for r in range(2, ws.max_row + 1):
    m_val = ws.cell(row=r, column=13).value
    code, reason = smart_match(m_val, d_map)

    if code == "歧义":
        ws.cell(row=r, column=14).value = reason
        stats["ambiguous"] += 1
    elif code is not None:
        ws.cell(row=r, column=14).value = code
        if reason is None:
            if m_val and str(m_val).strip() in d_map:
                stats["exact"] += 1
            else:
                stats["fuzzy"] += 1
        else:
            stats["fuzzy"] += 1
    else:
        ws.cell(row=r, column=14).value = reason
        if reason == "M列为空" or reason == "M列为null":
            stats["empty"] += 1
        else:
            stats["not_found"] += 1

wb.save(OUT)
print(f"保存成功: {OUT}")
print(f"统计: 精确匹配={stats['exact']}, 模糊匹配={stats['fuzzy']}, "
      f"歧义={stats['ambiguous']}, 未找到={stats['not_found']}, 空值={stats['empty']}")
print(f"总计处理行数: {ws.max_row - 1}")
