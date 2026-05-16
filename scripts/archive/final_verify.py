import openpyxl

wb = openpyxl.load_workbook(r"D:\MyDownloads\阿发_已匹配_v2.xlsx")
ws = wb["收货省市"]

print("=== 抽样验证 (别名匹配 & 各类匹配) ===\n")
print(f"{'Row':<6} {'M(数仓-市)':<38} {'N(城市编码)':<15} {'匹配方式'}")
print("-" * 80)

alias_samples = []
for r in range(2, ws.max_row + 1):
    m = ws.cell(row=r, column=13).value
    n = ws.cell(row=r, column=14).value
    if not m or not n:
        continue
    m_str = str(m).strip()
    n_str = str(n).strip() if n else ""
    if m_str in ["凉山州","伊犁州","恩施州","大理州","昌吉州","吐鲁番地区",
                 "哈密地区","日喀则地区","山南地区","那曲地区","海东地区",
                 "昌都地区","海北州","海西州","甘南州","黄南州","海南州",
                 "果洛州","玉树州","红河州","文山州","西双版纳州","德宏州",
                 "迪庆州","黔西南州","黔东南州","黔南州","临夏州",
                 "博尔塔拉州","巴音郭楞州","阿坝州","甘孜州","楚雄州",
                 "乐东县","保亭县","陵水县","琼中县","白沙县","昌江县",
                 "怒江州"]:
        if len(alias_samples) < 10:
            alias_samples.append((r, m_str, n_str))
    if len(alias_samples) >= 10:
        break

for r, m_str, n_str in alias_samples:
    print(f"{r:<6} {m_str:<38} {n_str:<15} 别名→标准名")

# 特殊原因行
print()
print("=== 特殊原因行 ===")
for r in range(2, ws.max_row + 1):
    n = ws.cell(row=r, column=14).value
    if n and isinstance(n, str) and not n.isdigit():
        m = ws.cell(row=r, column=13).value
        print(f"Row {r}: M=[{m}] → N=[{n}]")

# 再次验证所有编码格式
print()
print("=== 编码格式最终验证 ===")
bad = []
all_codes = set()
for r in range(2, ws.max_row + 1):
    n = ws.cell(row=r, column=14).value
    if n and isinstance(n, str) and n.isdigit():
        all_codes.add(n)
        if len(n) != 6:
            bad.append((r, n))

# 省份前缀检查
province_prefixes = {
    "11":"北京","12":"天津","13":"河北","14":"山西","15":"内蒙古",
    "21":"辽宁","22":"吉林","23":"黑龙江","31":"上海","32":"江苏",
    "33":"浙江","34":"安徽","35":"福建","36":"江西","37":"山东",
    "41":"河南","42":"湖北","43":"湖南","44":"广东","45":"广西",
    "46":"海南","50":"重庆","51":"四川","52":"贵州","53":"云南",
    "54":"西藏","61":"陕西","62":"甘肃","63":"青海","64":"宁夏",
    "65":"新疆","71":"台湾","81":"香港","82":"澳门",
}
for code in sorted(all_codes):
    prefix = code[:2]
    if prefix not in province_prefixes:
        bad.append((0, code))

if not bad:
    print(f"  ✓ 所有 {len(all_codes)} 个不同编码格式正确(6位数字,省份前缀有效)")
else:
    print(f"  ✗ 发现 {len(bad)} 个异常编码:")
    for r, c in bad:
        print(f"    Row {r}: {c}")

print(f"\n=== 最终覆盖统计 ===")
total = ws.max_row - 1
filled = 0
special = 0
empty = 0
for r in range(2, ws.max_row + 1):
    n = ws.cell(row=r, column=14).value
    if n is None:
        empty += 1
    elif isinstance(n, str) and n.isdigit():
        filled += 1
    else:
        special += 1
print(f"已填入编码: {filled} ({filled/total*100:.1f}%)")
print(f"特殊原因:   {special}")
print(f"空值:       {empty}")
print(f"总计行:     {total}")
