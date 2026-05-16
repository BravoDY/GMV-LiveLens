import openpyxl

wb = openpyxl.load_workbook(r"D:\MyDownloads\阿发.xlsx")
ws = wb["收货省市"]
print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
print()

print("=== Header Row (Row 1) ===")
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    letter = openpyxl.utils.get_column_letter(col)
    print(f"  Col {col} ({letter}): {val}")

print()
print("=== First 50 rows of C, D, M, N ===")
print(f"{'Row':<6} {'C':<40} {'D':<20} {'M':<40} {'N':<20}")
for r in range(1, min(51, ws.max_row + 1)):
    c = ws.cell(row=r, column=3).value
    d = ws.cell(row=r, column=4).value
    m = ws.cell(row=r, column=13).value
    n = ws.cell(row=r, column=14).value
    print(f"{r:<6} {str(c)[:38]:<40} {str(d)[:18]:<20} {str(m)[:38]:<40} {str(n)[:18]:<20}")
