import openpyxl

wb = openpyxl.load_workbook(r"D:\MyDownloads\阿发_已匹配.xlsx")
ws = wb["收货省市"]

print("=== 模糊匹配与歧义的行 ===")
for r in range(2, ws.max_row + 1):
    n = ws.cell(row=r, column=14).value
    if n and isinstance(n, str):
        m = ws.cell(row=r, column=13).value
        # 判断是否是编码(纯数字)还是原因文本
        if not n.isdigit():
            print(f"Row {r}: M=[{m}] → N=[{n}]")

print()

# 确认哪些已经在C列但match时走的精确匹配
d_map = {}
for r in range(2, ws.max_row + 1):
    c = ws.cell(row=r, column=3).value
    d = ws.cell(row=r, column=4).value
    if c and d:
        d_map[str(c).strip()] = str(d).strip()

# 统计匹配结果分布
exact = 0
fuzzy = 0
ambig = 0
not_found = 0
empty = 0
for r in range(2, ws.max_row + 1):
    m = ws.cell(row=r, column=13).value
    n = ws.cell(row=r, column=14).value
    if n is None:
        empty += 1
    elif isinstance(n, str) and not n.isdigit():
        if "歧义" in n or "候选" in n:
            ambig += 1
        else:
            not_found += 1
    elif isinstance(n, str) and n.isdigit():
        m_str = str(m).strip() if m else ""
        if m_str in d_map:
            exact += 1
        else:
            fuzzy += 1
    else:
        # numeric stored as int
        m_str = str(m).strip() if m else ""
        if m_str in d_map:
            exact += 1
        else:
            fuzzy += 1

print(f"精确匹配: {exact}")
print(f"模糊匹配: {fuzzy}")
print(f"歧义: {ambig}")
print(f"未找到: {not_found}")
print(f"空值: {empty}")
print(f"合计: {exact + fuzzy + ambig + not_found + empty}")
