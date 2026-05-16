import openpyxl
from collections import Counter

wb = openpyxl.load_workbook(r"D:\MyDownloads\阿发_已匹配.xlsx")
ws = wb["收货省市"]

unmatched = []
for r in range(2, ws.max_row + 1):
    m = ws.cell(row=r, column=13).value
    n = ws.cell(row=r, column=14).value
    if m and n and isinstance(n, str) and not n.isdigit():
        unmatched.append(str(m).strip())

counter = Counter(unmatched)
for city, cnt in counter.most_common():
    print(f"  [{city}] x{cnt}")

print(f"\n总计 {len(counter)} 种不同未匹配城市")
