import openpyxl

wb = openpyxl.load_workbook(r"D:\MyDownloads\阿发.xlsx")
ws = wb["收货省市"]

d_map = {}
for r in range(2, ws.max_row + 1):
    c = ws.cell(row=r, column=3).value
    d = ws.cell(row=r, column=4).value
    if c and d:
        d_map[str(c).strip()] = str(d).strip()

m_set = set()
for r in range(2, ws.max_row + 1):
    m = ws.cell(row=r, column=13).value
    if m:
        m_set.add(str(m).strip())

only_m = sorted(m_set - set(d_map.keys()))

print(f"C→D 编码映射共 {len(d_map)} 条")
print(f"M列中完全匹配: {len(m_set & set(d_map.keys()))} 个")
print(f"M列中未匹配: {len(only_m)} 个")
print()

print("=== C→D 完整映射 ===")
for name, code in sorted(d_map.items()):
    print(f"  [{name}] → {code}")

print()
print(f"=== 未匹配的M列城市 ({len(only_m)}) ===")
for city in only_m:
    print(f"  [{city}]")
